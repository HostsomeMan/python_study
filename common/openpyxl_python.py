#-*-coding:utf-8-*-

'''
· font(字体类)：字号、字体颜色、下划线等

· fill(填充类)：颜色等

· border(边框类)：设置单元格边框

· alignment(位置类)：对齐方式

· number_format(格式类)：数据格式

· protection(保护类)：写保护

'''

from openpyxl import Workbook
from openpyxl import load_workbook

import _datetime
import time
import os

# # 创建文件对象
wb = Workbook()

# # 获取第一个sheet
# ws = wb.active
#
# # 写入数字
# ws['A1'] = 43
# ws['A2'] = '测试'
#
# # 写入一个当前时间
# ws['A3'] = _datetime.datetime.now()
# # 写入一个自定义的时间格式
# ws['A4'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
# wb.save('E:/EXCEL_TEST.xlsx')

# 写入多个单元格

# ======================================================================================================================
# 创建sheet
# ws1 = wb.create_sheet("Mysheet")        # 创建一个sheet
# ws1.title = "New Title"                 # 设定一个sheet的名字
# ws2 = wb.create_sheet("Mysheet", 0)     # 设定sheet的插入位置，默认插在后面
# ws2.title = u"你好"                       # 设定一个sheet的名字，必须是Unicode
#
# # 获取全部sheet的名字，病例sheet名字
# print(wb.sheetnames)
# for sheet_name in wb.sheetnames:
#     print(sheet_name)
#
# print("*"*5)
#
# for sheet in wb:
#     print(sheet.title)
#
# # 复制一个sheet
# wb["New Title"]["A1"] = "zeke"
# source = wb["New Title"]
# target = wb.copy_worksheet(source)
#
# wb.save("e:\\sample.xlsx")

# ======================================================================================================================

# 操作单元格
# ws1 = wb.create_sheet("Mysheet")         # 创建一个sheet
#
# ws1['A1'] = 1.22
# ws1['B1'] = "你好"
# d = ws1.cell(row=4, column=2, value=10)     # 在第四行，第二列输入值10
#
# print(ws1["A1"].value)
# print(ws1["B1"].value)
# print(d.value)
#
# wb.save("e:\\sample1.xlsx")

# ======================================================================================================================

# 操作批量的单元格
# 无论ws.rows还是ws.iter_rows都是一个对象
# 除上述两个对象外，单行、单列都是一个元祖，多行多列是二维元祖
# ws1 = wb.create_sheet("Mysheet")
#
# ws1["A1"]=1
# ws1["A2"]=2
# ws1["A3"]=3
#
# ws1["B1"]=4
# ws1["B2"]=5
# ws1["B3"]=6
#
# ws1["C1"]=7
# ws1["C2"]=8
# ws1["C3"]=9
#
# # 操作单列
# print(ws1["A"])
# for cell in ws1["A"]:
#     print(cell.value)
#
# # 操作列，获取每一个值
# print(ws1["A:C"])
# for cloumn in ws1["A:C"]:
#     for cell in cloumn:
#         print(cell.value)
#
# #操作多行
# row_range = ws1[1:3]
# print(row_range)
# for row in row_range:
#     for cell in row:
#         print(cell.value)
#
# print("*"*50)
# for row in ws1.iter_rows(min_row=1, min_col=1, max_col=3, max_row=3):
#     for cell in row:
#         print(cell.value)
#
# #获取所有行，按行取值
# print(ws1.rows)
# for row in ws1.rows:
#     print(row)
#     for cell in row:
#         print(cell.value)
#
# print("*"*50)
# #获取所有列，按列取值
# print(ws1.columns)
# for col in ws1.columns:
#     print(col)
#
# wb.save("e:\\sample2.xlsx")

# ======================================================================================================================

# 使用百分数
from openpyxl import load_workbook
wb1 = load_workbook("e:\\sample.xlsx")
wb1.guess_types = True
ws=wb.active
ws["D1"]="12%"
print(ws["D1"].value)

# Save the file
wb.save("e:\\sample.xlsx")

# ======================================================================================================================
# 获取所有的行对象
# wb1 = load_workbook("e:\\sample3.xlsx")
# ws = wb1.active
# rows = []
