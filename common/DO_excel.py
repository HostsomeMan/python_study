import openpyxl,os
from openpyxl import load_workbook
from comm.get_project_path import Project_path
class  Do_excel:
    
    def __init__(self,sheet_name="Sheet1"):
        self.sheet_name=sheet_name
        self.excel_data_path=Project_path().project_path()+"test_data"+os.sep
    def read_excel(self,name):
        read_path=self.excel_data_path+name
        wb=load_workbook(read_path)
        sheet=wb[self.sheet_name]
        rows=sheet.max_row
        cols=sheet.max_column
        excel_data=[]
        for i in  range(2,rows+1):
            rows_data={}
            rows_data["url"]=sheet.cell(i,1).value
            rows_data["method"]=sheet.cell(i,2).value
            rows_data["params"]=sheet.cell(i,3).value
            rows_data["expect"]=sheet.cell(i,4).value
            rows_data["actull"]=sheet.cell(i,5).value
            rows_data["id"]=sheet.cell(i,7).value
            excel_data.append(rows_data)
        return excel_data
    def write_excel(self,row,col,value,name):
        path=self.excel_data_path+name
        wb=load_workbook(path)
        sheet=wb[self.sheet_name]        
        sheet.cell(row,col).value=value
        wb.save(path)
